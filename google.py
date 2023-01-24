import pandas as pd
import pygsheets
import openpyxl

from pygsheets import DataRange

gc = pygsheets.authorize(service_file=r"C:\Users\user\Desktop\stable-woods-374912-6149e61555af.json")



sh = gc.open_by_url('https://docs.google.com/spreadsheets/d/14clInBRzbD8RzxUqg7aXv-u2eeQw_DrLHaoEn2J8GHc/edit#gid=931961663g')

wks = sh[1]



#wks.set_dataframe(df, (1, 1))

color = {
#"blue": 1,
  "red": 2,
  "green": 23,
  "alpha": 1
}

wks2 = sh[0]

df = wks2.get_as_df()

l = df.shape[0]+1

gc.set_batch_mode(True)
model_cell = wks2.cell('A1')
model_cell.set_text_format('bold', True)
model_cell.text_format['fontSize'] = 11
model_cell.color = (255/255, 132/255, 0, 1)
model_cell.wrap_strategy = 'WRAP'
model_cell.borders = {
    "bottom": {'style': 'SOLID_THICK', "width": 1,"color": {'red': 0, 'green': 0, 'blue': 0}},
    "right": {'style': 'SOLID_THICK', "width": 1,"color": {'red': 0, 'green': 0, 'blue': 0}}
}


date_cell = wks2.cell('P1')
date_cell.set_number_format(format_type= pygsheets.FormatType.DATE)
date_cell.wrap_strategy = 'WRAP'
date_cell.borders = {
    "bottom": {'style': 'SOLID', "width": 1,"color": {'red': 0, 'green': 0, 'blue': 0}},
    "right": {'style': 'SOLID', "width": 1,"color": {'red': 0, 'green': 0, 'blue': 0}},
    "left": {'style': 'SOLID', "width": 1,"color": {'red': 0, 'green': 0, 'blue': 0}}
}




wks2.add_conditional_formatting('L', 'L', 'NUMBER_GREATER_THAN_EQ', {'backgroundColor': {'red': 0.8},
                                'textFormat': {'bold': True}}, ['30'])

wks2.add_conditional_formatting('L', 'L', 'NUMBER_BETWEEN', {'backgroundColor':
{"red": 250/255, "green": 77/255, "blue": 77/255, "alpha": 1}}, ['14', '30'])

wks2.add_conditional_formatting('L', 'L', 'NUMBER_LESS_THAN_EQ', {'backgroundColor':
{"red": 255/255, "green": 158/255, "blue": 158/255, "alpha": 1}}, ['14'])

border_cell = wks2.cell('R1')
border_cell.borders = {
    "bottom": {'style': 'SOLID', "width": 1,"color": {'red': 0, 'green': 0, 'blue': 0}},
    "right": {'style': 'SOLID', "width": 1,"color": {'red': 0, 'green': 0, 'blue': 0}},
    "left": {'style': 'SOLID', "width": 1,"color": {'red': 0, 'green': 0, 'blue': 0}}
}
border_cell.wrap_strategy = 'WRAP'
l = str(l)

l = 'L'+l



DataRange('A2',l, worksheet=wks2).apply_format(border_cell, fields="userEnteredFormat.borders") #fields="userEnteredFormat.wrap_strategy"
DataRange('A2',l, worksheet=wks2).apply_format(border_cell, fields="userEnteredFormat.wrap_strategy")
#DataRange('B','B', worksheet=wks2).apply_format(date_cell)

DataRange('A1','L1', worksheet=wks2).apply_format(model_cell)

gc.run_batch()

gc.set_batch_mode(False)



#
# wks2.add_conditional_formatting('N', 'N', 'NUMBER_BETWEEN', {'backgroundColor':
# {"red": 250/255, "green": 77/255, "blue": 77/255, "alpha": 1}}, ['14', '30'])
#
# wks2.add_conditional_formatting('N', 'N', 'NUMBER_LESS_THAN_EQ', {'backgroundColor':
# {"red": 255/255, "green": 158/255, "blue": 158/255, "alpha": 1}}, ['14'])
#
# wks2.adjust_column_width(start=1, end=1, pixel_size=30)
# #read = wks.get_as_df()
#
#
#
# wks2.add_conditional_formatting('K', 'K', 'TEXT_EQ', {'backgroundColor':color}, ['Выдан'])
#
# gc.run_batch()
#
# gc.set_batch_mode((False))
#
# model_cell = wks2.cell('A1')
# model_cell.set_text_format('bold', True)
# model_cell.text_format['fontSize'] = 11
# #model_cell.text_format['wrapStrategy'] = 'WRAP'
# model_cell.color = (0.5, 0.5, 0.5, 1)
# model_cell.wrap_strategy = 'WRAP'
# model_cell.borders ={"bottom": {'style': 'SOLID', "width": 1,"color": {'red':1}}}
#
#
# DataRange('A2','D2', worksheet=wks2).apply_format(model_cell)
#
# DataRange('E', 'E', worksheet=wks2).apply_format(model_cell)
#
#
# DataRange('A1', 'L1', worksheet=wks2).update_borders(top=True, right=True, bottom=True, left=True, style='SOLID_THICK', red=0, green=0, blue=0)